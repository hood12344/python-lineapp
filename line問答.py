# 技術文章
# https://developers.line.biz/en/docs/messaging-api/sending-messages/#methods-of-sending-message

# 一定要改的地方～～
# 請打開 ine.xlsx 中的  setup 改 auth_token 	接收人的id
# auth_token="2bSr7EUf1Ip1OVYBgh7HgCpdrPj/KyphlQQpQKkyUN3NVuZlQ36dp2b3O6imlhLk2LunCyORmlfQHb9tNXaul716CTlDNSD78PPUaL3ZM51a5yvLf6oWBQ4zyeBcAgD8Agc478HreI5bD5ArdslHrwdB04t89/1O/w1cDnyilFU="
#接收人的id=["U22869b7bf55a2026578867d615fe8c11"]

import requests
import json
from openpyxl import Workbook     # pip install openpyxl
import time                       # 時間
from openpyxl import load_workbook


wb = load_workbook('line.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheetSetup = wb["setup"]         # 打開一個工作欄
auth_token=sheetSetup.cell(row=2, column=1).value
接收人的id=sheetSetup.cell(row=2, column=2).value


# 請參考
# https://developers.line.biz/zh-hant/docs/messaging-api/sending-messages/#methods-of-sending-message
message = {
    "to": [接收人的id],
    "messages": [
                {
                    "type": "text",
                    "text": "1. 你好 \n  5ac21184040ab15980c9b43a150"
                }
            ]
    }

# 資料回傳 到 Line 的 https 伺服器
hed = {'Authorization': 'Bearer ' + auth_token}
url = 'https://api.line.me/v2/bot/message/multicast'

requests.post(url, json=message, headers=hed)      # 把資料HTTP POST送出去



"""
curl -v -X POST https://api.line.me/v2/bot/message/multicast \
-H 'Content-Type: application/json' \
-H 'Authorization: Bearer {channel access token}' \
-d '{
    "to": ["U4af4980629...","U0c229f96c4..."],
    "messages":[
        {
            "type":"text",
            "text":"Hello, world1"
        },
        {
            "type":"text",
            "text":"Hello, world2"
        }
    ]
}'
"""