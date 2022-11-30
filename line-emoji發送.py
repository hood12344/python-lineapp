# 技術文章
# https://developers.line.biz/en/docs/messaging-api/sending-messages/#methods-of-sending-message

# 一定要改的地方～～
# 請打開 ine.xlsx 中的  setup 改 auth_token 	接收人的id
#auth_token="R4XTGnlFzJ+hYSTA8mp47OyRGQbPR5hzXdAYDPaZXBsz7UDBQCHFeXLPpoG2WWsYfGq3HKAOrslaiAlKqix8+k2bk9xaOusBZKzLJMvJimMXEE6heI5Haou+EnW3sfRYow4QEK16p6ikuVL7y3aX2AdB04t89/1O/w1cDnyilFU="
#接收人的id=["U5f2b01b7ea1b9c122e7d4c7bb654be76"]

import requests
import json
from openpyxl import Workbook  # pip install openpyxl
import time  # 時間
from openpyxl import load_workbook
from sys import version as python_version
from cgi import parse_header, parse_multipart
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
from urllib.parse import parse_qs

wb = load_workbook('line.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheetSetup = wb["setup"]  # 打開一個工作欄
auth_token = sheetSetup.cell(row=2, column=1).value
接收人的id=sheetSetup.cell(row=2, column=2).value

# 接收人的id=sheetSetup.cell(row=2, column=2).value


class MyHandler(RequestHandler):
    def do_POST(self):
        varLen = int(self.headers['Content-Length'])  # 取得讀取進來的網路資料長度
        if varLen > 0:
            post_data = self.rfile.read(varLen)  # 讀取傳過來的資料
            data = json.loads(post_data)  # 把字串 轉成JSON
            print(data)
            replyToken = data['events'][0]['replyToken']  # 回傳要用Token
            userId3 = data['events'][0]['source']['userId']  # 傳資料過來的使用者是誰
            text = data['events'][0]['message']['text']  # 用戶的傳遞過來的文字內容
            傳過來的資料型態 = data['events'][0]['message']['type']  # 傳過來的資料型態
            # str1="$ and $"
            # index1=str1.find("$")
        # 請參考
        # Text message example with LINE emoji
        # 所有emojis 列表
        # https://developers.line.biz/en/docs/messaging-api/emoji-list/#specify-emojis-in-message-object

        message = {
            "replyToken": replyToken,
            "messages": [
                {
                    "type": "text",
                    "text": "1. 你的LineID是: " + userId3 + "\n  傳過來的文字是:" + text
                },
                {
                    "type": "text",
                    "text": "$ LINE emoji $",
                    "emojis": [
                        {
                            "index": 0,
                            "productId": "5ac221ca040ab15980c9b449",
                            "emojiId": "004"
                        },
                        {
                            "index": 13,
                            "productId": "5ac1bfd5040ab15980c9b435",
                            "emojiId": "002"
                        }
                    ]
                }
            ]
        }

        # 資料回傳 到 Line 的 https 伺服器
        hed = {'Authorization': 'Bearer ' + auth_token}
        url = 'https://api.line.me/v2/bot/message/reply'
        self.send_response(200)
        self.end_headers()
        requests.post(url, json=message, headers=hed)  # 把資料HTTP POST送出去


socketserver.TCPServer.allow_reuse_address = True  # 可以重複使用IP
httpd = socketserver.TCPServer(('0.0.0.0', 8000), MyHandler)  # 啟動WebServer   :8888
try:
    httpd.serve_forever()  # 等待用戶使用 WebServer
except:
    print("Closing the server.")
    httpd.server_close()  # 關閉 WebServer

"""
{
"destination":"Uc56db95c5bdabf2405a3c1c5afb5466d",
"events":[
   {"type":"message",
    "message":{
        "type":"text",
        "id":"16026143839698",
        "text":"111"},
    "webhookEventId":"01G26BHBNVMG06ZSEE27G4GJDB",
    "deliveryContext":{"isRedelivery":false},
    "timestamp":1651628355014,
    "source":{
        "type":"user",
        "userId":"U22869b7bf55a2026578867d615fe8c11"},
        "replyToken":"d35b62f1927d42b0ba4523d79e92e84e",
        "mode":"active"}
    ]
}'
"""

"""
所有emojis 列表
https://developers.line.biz/en/docs/messaging-api/emoji-list/#specify-emojis-in-message-object
message = {
            "replyToken":replyToken,
            "messages": [
                {
                    "type": "text",
                    "text": "1. 你的LineID是: "+userId3+"\n  傳過來的文字是:"+text
                },
                // Text message example with LINE emoji
                {
                    "type": "text",
                    "text": "$ LINE emoji $",
                    "emojis": [
                      {
                        "index": 0,
                        "productId": "5ac221ca040ab15980c9b449",
                        "emojiId": "004"
                      },
                      {
                        "index": 13,
                        "productId": "5ac1bfd5040ab15980c9b435",
                        "emojiId": "002"
                      }
                    ]
                }
            ]
        }





"""